import imaplib
import email
import os
import time
from openpyxl import load_workbook
from plyer import notification
import ollama

# ---------------- CONFIG ----------------
EMAIL = "EMail Address"      # Your Gmail Address
PASSWORD = "<PASSWORD>"  # Use Gmail App Password
IMAP_SERVER = "imap.gmail.com"
YOUR_ID = "<YOUR_ID>"               # Your ID
CHECK_INTERVAL = 120             # seconds

ATTACH_DIR = "attachments"
os.makedirs(ATTACH_DIR, exist_ok=True)
# ----------------------------------------


def notify(msg):
    notification.notify(
        title="ID Watcher",
        message=msg,
        timeout=8
    )
def id_in_mail_text(subject, body):
    combined = subject + " " + body
    return YOUR_ID in combined

def check_excel_with_llm(path):
    try:
        wb = load_workbook(path)
        text = ""
        found_exact = False

        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        val = str(cell)
                        text += val + " | "
                        if YOUR_ID in val:
                            found_exact = True

        # If exact match found, no need to ask LLM
        
        if found_exact:
            print("Exact ID match found in Excel")
            return True

        prompt = f"""
My ID is: {YOUR_ID}

Below is spreadsheet content:
{text}

Question:
Does this file contain my ID or clearly refer to me?
Reply with only YES or NO.
"""

        response = ollama.chat(
            model="llama3",
            messages=[{"role": "user", "content": prompt}]
        )

        answer = response["message"]["content"].strip().upper()
        print("LLM Answer:", answer)

        return "YES" in answer

    except Exception as e:
        print("Excel/LLM error:", e)
        return False

from datetime import datetime
def get_mail_body(msg):
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain" and not part.get("Content-Disposition"):
                body += part.get_payload(decode=True).decode(errors="ignore")
    else:
        body = msg.get_payload(decode=True).decode(errors="ignore")
    return body.strip()
def extract_company_with_llm(subject, body):
    prompt = f"""
This email is from a placement cell.

Subject:
{subject}

Body:
{body}

Question:
Which company is this email about?
Reply with only the company name. If unknown, reply UNKNOWN.
"""

    response = ollama.chat(
        model="llama3",
        messages=[{"role": "user", "content": prompt}]
    )

    return response["message"]["content"].strip()
from datetime import datetime

def check_mail():
    print("Checking today's mail...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL, PASSWORD)
    mail.select("inbox")

    today = datetime.now().strftime("%d-%b-%Y")
    status, data = mail.search(
    None,
    f'(UNSEEN SINCE "{today}" OR FROM "vitianscdc2026@vitstudent.ac.in" FROM "test@gmail.com")'
)




    if not data or not data[0]:
        mail.logout()
        return

    for num in data[0].split():
        _, msg_data = mail.fetch(num, '(RFC822)')

        msg = email.message_from_bytes(msg_data[0][1])

        subject = msg.get("Subject", "")
        body = get_mail_body(msg)

        company = extract_company_with_llm(subject, body)

        # 1) Check ID in mail text
        if id_in_mail_text(subject, body):
            notify(
                f"ID found in MAIL!\nCompany: {company}\nSubject: {subject}"
            )

        # 2) Check Excel attachments
        for part in msg.walk():
            if part.get_content_disposition() == "attachment":
                filename = part.get_filename()
                if filename and filename.lower().endswith(".xlsx"):
                    path = os.path.join(ATTACH_DIR, filename)
                    with open(path, "wb") as f:
                        f.write(part.get_payload(decode=True))

                    print(f"Downloaded: {filename}")

                    if check_excel_with_llm(path):
                        notify(
                            f"ID found in EXCEL!\nCompany: {company}\nSubject: {subject}"
                        )

    mail.logout()


if __name__ == "__main__":
    notify("ID Watcher started")
    while True:
        try:
            check_mail()
        except Exception as e:
            print("Mail error:", e)

        time.sleep(CHECK_INTERVAL)
